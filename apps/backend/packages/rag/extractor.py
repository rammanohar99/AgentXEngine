"""
Document extractor — converts raw file bytes into plain text for the RAG pipeline.

Supported formats:
  PDF   (.pdf)              — pypdf text extraction + Gemini Vision fallback for scanned PDFs
  Excel (.xlsx, .xls)       — openpyxl, all sheets converted to readable text tables
  CSV   (.csv)              — stdlib csv, formatted as aligned text table
  Image (.png, .jpg, .jpeg,
         .webp, .tiff, .bmp) — Gemini Vision API (no Tesseract system dependency)
  Text  (everything else)   — UTF-8 decode

Design:
- All extractors are pure functions: bytes → str
- Image extraction is async (Gemini API call); all others are sync
- Each extractor preserves structure: tables stay tabular, pages are separated
- Extraction errors raise ValueError with a human-readable message
- No global state — extractors are stateless utility functions
"""

from __future__ import annotations

import csv
import io
from typing import Any

import structlog

logger = structlog.get_logger(__name__)

# File extension → source_type label used in DocumentMetadata
EXTENSION_TO_SOURCE_TYPE: dict[str, str] = {
    ".pdf": "pdf",
    ".xlsx": "excel",
    ".xls": "excel",
    ".csv": "csv",
    ".png": "image",
    ".jpg": "image",
    ".jpeg": "image",
    ".webp": "image",
    ".tiff": "image",
    ".tif": "image",
    ".bmp": "image",
}

# Extensions that must go through the upload endpoint (binary / non-UTF-8)
BINARY_EXTENSIONS = set(EXTENSION_TO_SOURCE_TYPE.keys())


def get_source_type(filename: str) -> str:
    """Return the source_type label for a filename."""
    ext = _ext(filename)
    return EXTENSION_TO_SOURCE_TYPE.get(ext, "text")


def is_binary(filename: str) -> bool:
    """Return True if the file must be processed server-side (not read as text in browser)."""
    return _ext(filename) in BINARY_EXTENSIONS


def extract_text(content_bytes: bytes, filename: str) -> str:
    """
    Synchronous extraction for non-image formats.

    Raises ValueError if no text can be extracted.
    Use extract_image_text() for image files.
    """
    ext = _ext(filename)

    if ext == ".pdf":
        return _extract_pdf(content_bytes, filename)
    if ext in (".xlsx", ".xls"):
        return _extract_excel(content_bytes, filename)
    if ext == ".csv":
        return _extract_csv(content_bytes, filename)
    if ext in BINARY_EXTENSIONS:
        raise ValueError(
            f"'{filename}' is an image file. Use extract_image_text() for image extraction."
        )

    # Plain text fallback
    return content_bytes.decode("utf-8", errors="replace")


async def extract_image_text(
    content_bytes: bytes,
    filename: str,
    gemini_client: Any,
    model: str = "gemini-2.0-flash",
) -> str:
    """
    Extract text from an image using Gemini Vision API.

    Uses the already-configured google-genai client — no Tesseract required.
    Returns the extracted text, or raises ValueError if extraction fails.
    """

    from google.genai import types

    ext = _ext(filename)
    mime_map = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".tiff": "image/tiff",
        ".tif": "image/tiff",
        ".bmp": "image/bmp",
    }
    mime_type = mime_map.get(ext, "image/png")

    logger.info("image_ocr_start", filename=filename, mime_type=mime_type)

    prompt = (
        "Extract ALL text visible in this image verbatim. "
        "Preserve tables, lists, and structure as plain text. "
        "If the image contains a table, format it with | separators. "
        "If there is no text, respond with: (no text found)"
    )

    response = await gemini_client.aio.models.generate_content(
        model=model,
        contents=[
            types.Content(
                role="user",
                parts=[
                    types.Part(
                        inline_data=types.Blob(
                            mime_type=mime_type,
                            data=content_bytes,
                        )
                    ),
                    types.Part(text=prompt),
                ],
            )
        ],
    )

    text = response.text or ""
    if not text.strip() or text.strip() == "(no text found)":
        raise ValueError(f"No text could be extracted from image '{filename}'.")

    logger.info("image_ocr_complete", filename=filename, chars=len(text))
    return text


# ── Private extraction helpers ────────────────────────────────────────────────


def _ext(filename: str) -> str:
    """Return the lowercase file extension including the dot."""
    idx = filename.rfind(".")
    return filename[idx:].lower() if idx != -1 else ""


def _extract_pdf(content_bytes: bytes, filename: str) -> str:
    """Extract text from a PDF using pypdf."""
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("pypdf is not installed.") from exc

    reader = PdfReader(io.BytesIO(content_bytes))
    pages: list[str] = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"[Page {i + 1}]\n{text.strip()}")

    if not pages:
        raise ValueError(
            f"No extractable text found in '{filename}'. "
            "The PDF may be image-based — try uploading as an image instead."
        )

    return "\n\n".join(pages)


def _extract_excel(content_bytes: bytes, filename: str) -> str:
    """
    Extract text from an Excel workbook (.xlsx / .xls).

    Each sheet is rendered as a plain-text table with | separators.
    Sheet names are included as section headers.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed.") from exc

    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []

        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue
            rows.append([_cell_str(cell) for cell in row])

        if not rows:
            continue

        # Build column widths for alignment
        col_widths = [0] * max(len(r) for r in rows)
        for row in rows:
            for i, cell in enumerate(row):
                if i < len(col_widths):
                    col_widths[i] = max(col_widths[i], len(cell))

        # Format as aligned table
        lines: list[str] = [f"## Sheet: {sheet_name}"]
        for row_idx, row in enumerate(rows):
            padded = [cell.ljust(col_widths[i]) for i, cell in enumerate(row)]
            lines.append(" | ".join(padded).rstrip())
            # Add separator after header row
            if row_idx == 0:
                lines.append("-" * sum(col_widths) + "-" * (len(col_widths) * 3))

        sections.append("\n".join(lines))

    wb.close()

    if not sections:
        raise ValueError(f"No data found in Excel file '{filename}'.")

    return "\n\n".join(sections)


def _extract_csv(content_bytes: bytes, filename: str) -> str:
    """
    Extract text from a CSV file.

    Renders as a plain-text table with | separators.
    Tries UTF-8 first, falls back to latin-1.
    """
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"Could not decode CSV file '{filename}'.")

    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader if any(cell.strip() for cell in row)]

    if not rows:
        raise ValueError(f"No data found in CSV file '{filename}'.")

    # Build column widths
    col_count = max(len(r) for r in rows)
    col_widths = [0] * col_count
    for row in rows:
        for i, cell in enumerate(row):
            if i < col_count:
                col_widths[i] = max(col_widths[i], len(cell))

    lines: list[str] = [f"## {filename}"]
    for row_idx, row in enumerate(rows):
        padded = [
            row[i].ljust(col_widths[i]) if i < len(row) else " " * col_widths[i]
            for i in range(col_count)
        ]
        lines.append(" | ".join(padded).rstrip())
        if row_idx == 0:
            lines.append("-" * (sum(col_widths) + col_count * 3))

    return "\n".join(lines)


def _cell_str(value: Any) -> str:
    """Convert an Excel cell value to a clean string."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Avoid scientific notation for large integers stored as float
        if value == int(value):
            return str(int(value))
        return f"{value:.4g}"
    return str(value).strip()
